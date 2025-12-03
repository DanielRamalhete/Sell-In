
# src/process_excels.py
import os
import tempfile
from typing import Dict, List, Tuple

from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.client_credential import ClientCredential
from office365.sharepoint.files.file import File
from office365.sharepoint.folders.folder import Folder

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# --- Config via env vars (defina em GitHub Secrets / env) ---
SITE_URL         = "https://braveperspective.sharepoint.com/sites/equipa.comite"        # p.ex.: https://contoso.sharepoint.com/sites/Planeamento
FOLDER_PATHS_CSV = "/Documentos%20Partilhados/General/Teste%20-%20Daniel%20PowerAutomate/5.%20Planos%20Anuais/FMENEZES,/Documentos%20Partilhados/General/Teste%20-%20Daniel%20PowerAutomate/5.%20Planos%20Anuais/GMALAFAYA,/Documentos%20Partilhados/General/Teste%20-%20Daniel%20PowerAutomate/5.%20Planos%20Anuais/JPIRES,/Documentos%20Partilhados/General/Teste%20-%20Daniel%20PowerAutomate/5.%20Planos%20Anuais/TNAIA"
# 4 paths separados por vírgula (server-relative)
CLIENT_ID        = os.getenv("CLIENT_ID")
CLIENT_SECRET    = os.getenv("CLIENT_SECRET")

SHEET_NAME       = os.getenv("SOURCE_SHEET", "Resumo Plano anual")
NEW_SHEET_NAME   = os.getenv("TARGET_SHEET", "PowerBI nao mexer")

# Se quiser "dry run" sem upload, defina DRY_RUN=true
DRY_RUN = os.getenv("DRY_RUN", "false").lower() == "true"

# Cabeçalhos esperados
VALUE_COLS   = ["4Q", "1Q", "2Q", "3Q", "FY"]
PERCENT_COLS = ["4Q%", "1Q%", "2Q%", "3Q%", "FY%"]


def connect_ctx() -> ClientContext:
    # App Registration (client credentials)
    cred = ClientCredential(CLIENT_ID, CLIENT_SECRET)
    ctx = ClientContext(SITE_URL).with_credentials(cred)
    return ctx


def get_folder(ctx: ClientContext, server_relative_path: str) -> Folder:
    """
    server_relative_path exemplo:
    /sites/Planeamento/Documentos Partilhados/Pasta1
    """
    folder = ctx.web.get_folder_by_server_relative_url(server_relative_path)
    ctx.load(folder)
    ctx.execute_query()
    return folder


def download_file(ctx: ClientContext, folder: Folder, name: str, local_path: str):
    server_rel_url = f"{folder.serverRelativeUrl}/{name}"
    file = File(ctx.web, server_rel_url)
    with open(local_path, "wb") as f:
        file.download(f).execute_query()


def upload_file_replace(folder: Folder, local_path: str, name: str):
    """
    Simplesmente chama upload para o mesmo nome — o SharePoint trata o versionamento/overwrite.
    Para > ~250MB, usar upload em chunks; aqui esperamos ficheiros de planeamento < 50MB. [5](https://www.wmstephenscott.com/uploading-large-files-to-sharepoint-using-python/)
    """
    with open(local_path, "rb") as f:
        uploaded = folder.files.upload(name, f.read()).execute_query()
    return uploaded


def find_header_row(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    """
    Procura a linha onde está 'Marcas' e captura o índice de cada coluna relevante.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    # varrer primeiras ~50 linhas à procura dos cabeçalhos
    for r in range(1, min(max_row, 50) + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if values is None:
            continue
        # normalizar para str
        norm = [str(v).strip() if v is not None else "" for v in values]
        if "Marcas" in norm:
            # mapear nomes -> índices de coluna
            col_map = {}
            for name in ["Marcas"] + VALUE_COLS + PERCENT_COLS:
                if name in norm:
                    col_map[name] = norm.index(name) + 1  # coluna 1-based
            # validar colunas mínimas
            missing = [n for n in ["Marcas"] + VALUE_COLS + PERCENT_COLS if n not in col_map]
            if missing:
                raise ValueError(f"Faltam cabeçalhos na folha '{ws.title}': {missing}")
            return r, col_map

    raise ValueError("Não encontrei a linha de cabeçalhos (Marcas) na folha.")


def build_rows(ws: Worksheet, header_row: int, col_map: Dict[str, int]) -> List[List]:
    """
    Implementa a regra:
    - Linha com Marcas em branco => valores
    - Linha seguinte com Marcas preenchida => percentuais + nome da marca
    -> Gera uma única linha por marca: [Marca, 4Q,1Q,2Q,3Q,FY, 4Q%,1Q%,2Q%,3Q%,FY%]
    """
    rows_out: List[List] = []
    r = header_row + 1
    max_row = ws.max_row

    def val(row, col_name):
        c = ws.cell(row=row, column=col_map[col_name]).value
        return c

    while r <= max_row:
        marca_current = val(r, "Marcas")
        if marca_current is None or str(marca_current).strip() == "":
            # valores nesta linha
            marca_next = val(r + 1, "Marcas") if r + 1 <= max_row else None
            if marca_next is None or str(marca_next).strip() == "":
                # não é um par válido -> avançar
                r += 1
                continue

            values = [val(r, c) for c in VALUE_COLS]
            percents = [val(r + 1, c) for c in PERCENT_COLS]

            # normalizar percentuais (podem vir como '10%' texto; convertemos para número ou mantemos texto)
            def norm_percent(x):
                if isinstance(x, str) and x.endswith("%"):
                    try:
                        return float(x[:-1]) / 100.0
                    except:
                        return x
                return x

            percents = [norm_percent(p) for p in percents]

            rows_out.append([str(marca_next).strip()] + values + percents)

            r += 2  # saltar o par (valores + percentuais)
        else:
            # linha com marca preenchida sem par anterior — ignora, porque percentuais tratamos no par acima
            r += 1

    return rows_out


def write_new_sheet(wb: Workbook, name: str, rows: List[List]):
    if name in wb.sheetnames:
        ws = wb[name]
        # limpar
        wb.remove(ws)
    ws = wb.create_sheet(title=name)

    header = ["Marcas"] + VALUE_COLS + PERCENT_COLS
    ws.append(header)
    for row in rows:
        ws.append(row)


def process_workbook(local_path: str):
    wb = load_workbook(local_path, data_only=True)  # ler valores calculados (não fórmulas) [2](https://openpyxl.pages.heptapod.net/openpyxl/api/openpyxl.worksheet.worksheet.html)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Folha '{SHEET_NAME}' não existe em {os.path.basename(local_path)}")

    ws = wb[SHEET_NAME]
    header_row, col_map = find_header_row(ws)
    rows_out = build_rows(ws, header_row, col_map)
    write_new_sheet(wb, NEW_SHEET_NAME, rows_out)
    wb.save(local_path)  # grava alterações no ficheiro local
    return len(rows_out)


def main():
    ctx = connect_ctx()

    folder_paths = [p.strip() for p in FOLDER_PATHS_CSV.split(",") if p.strip()]
    if not folder_paths:
        raise RuntimeError("SP_FOLDER_PATHS_CSV está vazio.")

    for fp in folder_paths:
        folder = get_folder(ctx, fp)
        files = folder.files.get().execute_query()
        for f in files:
            name = f.name
            if not name.lower().endswith(".xlsx"):
                continue

            print(f"▶ Processar: {folder.serverRelativeUrl}/{name}")
            with tempfile.TemporaryDirectory() as tmp:
                local = os.path.join(tmp, name)

                # download
                download_file(ctx, folder, name, local)

                # processar
                n = process_workbook(local)
